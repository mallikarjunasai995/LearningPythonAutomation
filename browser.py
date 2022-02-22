


from re import M
from time import sleep
import webbrowser
from openpyxl import load_workbook
path = r'C:\\IOCVFDL_CC_I0P_OGT_Feb2.xlsx'

wb = load_workbook(path)  # Work Book
ws = wb.get_sheet_by_name('IOCVFDL')  # Work Sheet
column = ws['M']  # Column
print(column)
column_list = [column[x].value for x in range(len(column))]
print(column_list)
# webbrowser.open_new('http://10.207.33.48:8080/xplorer/console/ui')
# for i in column_list:
#     url = 'http://10.207.33.48:8080/xplorer/console/ui/' + i
#     webbrowser.open_new(url)
#     sleep(0.25)
   



# 1. launch ID's to map the xplorer session 
# 2. today we need to trigger IOP 
#


